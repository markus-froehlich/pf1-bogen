import openpyxl
path = "/Users/froema/Documents/Rollenspiel/Pathfinder/Pathinder Web App/Version 6.61/Bogen 6.61 Spieler.xlsx"
wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
print("Sheets total:", len(wb.sheetnames))
print(f"{'#':>2} {'state':<10} {'maxRow':>7} {'maxCol':>7}  name")
for i, ws in enumerate(wb.worksheets):
    print(f"{i:>2} {ws.sheet_state:<10} {ws.max_row:>7} {ws.max_column:>7}  {ws.title!r}")
print("\n=== Defined names ===")
try:
    names = list(wb.defined_names.keys())
    print("count:", len(names))
    for n in names:
        print("  ", n)
except Exception as e:
    print("err", e)
