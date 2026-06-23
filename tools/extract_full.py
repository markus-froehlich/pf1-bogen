"""Phase 2 foundation: COMPLETE untrimmed per-cell dump (constants + formulas + cached value).
This is the canonical data source (fixes the trimming caveat). One JSONL line per
non-empty cell: {"c": coord, "t": "f"|"v", "raw": formula-or-constant, "val": cached}.
"""
import openpyxl, json, os, warnings, datetime
warnings.simplefilter("ignore")
SRC = "/Users/froema/Documents/Rollenspiel/Pathfinder/Pathinder Web App/Version 6.61/Bogen 6.61 Spieler.xlsx"
OUT = "/Users/froema/Documents/Rollenspiel/Pathfinder/Pathinder Web App/extraction/sheets_full"
os.makedirs(OUT, exist_ok=True)

def norm(v):
    if isinstance(v, datetime.datetime):
        return v.isoformat()
    return v

print("loading formula wb...")
wb_f = openpyxl.load_workbook(SRC, data_only=False)
print("loading value wb (read_only)...")
wb_v = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

summary = []
for ws_f in wb_f.worksheets:
    title = ws_f.title
    ws_v = wb_v[title]
    # build cached-value dict by streaming the value sheet
    valmap = {}
    for row in ws_v.iter_rows():
        for c in row:
            if c.value is not None:
                valmap[c.coordinate] = c.value
    safe = title.replace("/", "_").replace(" ", "_")
    nconst = nform = 0
    with open(f"{OUT}/{safe}.jsonl", "w") as f:
        for row in ws_f.iter_rows():
            for c in row:
                v = c.value
                if v is None:
                    continue
                if isinstance(v, str) and v.startswith("="):
                    nform += 1
                    f.write(json.dumps({"c": c.coordinate, "t": "f", "raw": v,
                                        "val": norm(valmap.get(c.coordinate))}, ensure_ascii=False) + "\n")
                else:
                    nconst += 1
                    f.write(json.dumps({"c": c.coordinate, "t": "v", "raw": norm(v),
                                        "val": norm(valmap.get(c.coordinate, v))}, ensure_ascii=False) + "\n")
    summary.append({"sheet": title, "const": nconst, "formula": nform, "total": nconst + nform})
    print(f"  {title:<22} const={nconst:>7} formula={nform:>7}")

json.dump(summary, open(f"{OUT}/_summary.json", "w"), ensure_ascii=False, indent=2)
print("total cells:", sum(s["total"] for s in summary))
print("DONE full dump")
