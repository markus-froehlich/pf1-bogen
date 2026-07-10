"""Phase 1a: structure extraction (formulas, named ranges, validations, merged cells).
Loads workbook with formulas (data_only=False)."""
import openpyxl, json, os, csv, warnings
warnings.simplefilter("ignore")

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(_ROOT, "Version 6.61", "Bogen 6.61 Spieler.xlsx")
OUT = os.path.join(_ROOT, "extraction")

print("loading (formulas)...")
wb = openpyxl.load_workbook(SRC, data_only=False)

# ---- overview ----
overview = []
for ws in wb.worksheets:
    dv = ws.data_validations.dataValidation if ws.data_validations else []
    overview.append({
        "name": ws.title,
        "state": ws.sheet_state,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "merged": len(ws.merged_cells.ranges),
        "validations": len(dv),
    })
json.dump(overview, open(f"{OUT}/workbook_overview.json","w"), ensure_ascii=False, indent=2)
print("sheets:", len(overview))

# ---- named ranges ----
named = {}
for name, d in wb.defined_names.items():
    entries = []
    try:
        for sheet, coords in d.destinations:
            entry = {"sheet": sheet, "coords": coords}
            # single cell? resolve formula/constant
            cc = coords.replace("$","")
            if ":" not in cc:
                try:
                    cell = wb[sheet][cc]
                    entry["formula"] = cell.value
                except Exception as e:
                    entry["err"] = str(e)
            entries.append(entry)
    except Exception as e:
        entries.append({"raw": str(d.value), "err": str(e)})
    named[name] = entries
json.dump(named, open(f"{OUT}/named_ranges.json","w"), ensure_ascii=False, indent=2)
print("named ranges:", len(named))

# ---- data validations (dropdowns = input model) ----
vals = {}
for ws in wb.worksheets:
    if not ws.data_validations: continue
    lst = []
    for dv in ws.data_validations.dataValidation:
        lst.append({
            "type": dv.type,
            "formula1": dv.formula1,
            "formula2": dv.formula2,
            "sqref": str(dv.sqref),
            "allow_blank": dv.allow_blank,
            "prompt": dv.prompt,
        })
    if lst: vals[ws.title] = lst
json.dump(vals, open(f"{OUT}/validations.json","w"), ensure_ascii=False, indent=2)
print("sheets with validations:", len(vals))

# ---- merged cells ----
merged = {ws.title: [str(r) for r in ws.merged_cells.ranges] for ws in wb.worksheets if ws.merged_cells.ranges}
json.dump(merged, open(f"{OUT}/merged_cells.json","w"), ensure_ascii=False, indent=2)

# ---- per-sheet formulas dump (only formula cells) ----
os.makedirs(f"{OUT}/sheets_formulas", exist_ok=True)
total_f = 0
for ws in wb.worksheets:
    rows = []
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, str) and v.startswith("="):
                rows.append((c.coordinate, v))
    total_f += len(rows)
    safe = ws.title.replace("/","_").replace(" ","_")
    with open(f"{OUT}/sheets_formulas/{safe}.tsv","w",newline="") as f:
        w = csv.writer(f, delimiter="\t")
        w.writerow(["cell","formula"])
        w.writerows(rows)
print("total formula cells dumped:", total_f)
print("DONE structure")
